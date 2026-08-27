# -*- coding: utf-8 -*-
"""
Exports the 10,000-record synthetic transaction dataset (results.json,
which already has the Stage A risk scores + Stage B Gemini reasoning
for flagged records) into a CSV and an XLSX file, ready to upload into
Google Sheets or open directly in Excel.

Run: python export_to_spreadsheet.py
Output: fraud_transactions.csv, fraud_transactions.xlsx
"""
import json
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

with open("results.json", encoding="utf-8") as f:
    records = json.load(f)

rows = []
for r in records:
    rows.append({
        "transaction_id": r["transaction_id"],
        "account_id": r["account_id"],
        "customer_name": r.get("customer_name", ""),
        "address": r.get("address", ""),
        "account_type": r.get("account_type", ""),
        "timestamp": r["timestamp"],
        "amount": r["amount"],
        "channel": r["channel"],
        "merchant": r["merchant"],
        "device_id": r["device_id"],
        "ip_country": r["ip_country"],
        "ip_city": r["ip_city"],
        "customer_tenure_days": r["customer_tenure_days"],
        "prior_avg_txn_amount": r["prior_avg_txn_amount"],
        "risk_score": r.get("risk_score", 0),
        "flagged": r.get("flagged", False),
        "signals": "; ".join(r.get("signals", [])),
        "recommended_action": r.get("recommended_action", ""),
        "confidence": r.get("confidence", ""),
        "reasoning": r.get("reasoning", ""),
        "injected_fraud_pattern": r.get("injected_fraud_pattern") or "",
    })

df = pd.DataFrame(rows)

csv_path = "fraud_transactions.csv"
df.to_csv(csv_path, index=False, encoding="utf-8-sig")  # utf-8-sig so Excel/Sheets handle it cleanly
print(f"Wrote {csv_path} ({len(df)} rows)")

xlsx_path = "fraud_transactions.xlsx"
df.to_excel(xlsx_path, index=False, sheet_name="Transactions")

# Light formatting pass: bold header, freeze header row, autofit-ish column
# widths, highlight flagged rows.
from openpyxl import load_workbook
wb = load_workbook(xlsx_path)
ws = wb["Transactions"]

header_font = Font(bold=True, color="FFFFFF", name="Arial")
header_fill = PatternFill("solid", fgColor="4A1420")
for cell in ws[1]:
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal="center")

ws.freeze_panes = "A2"

widths = {
    "A": 15, "B": 13, "C": 20, "D": 32, "E": 11,
    "F": 22, "G": 12, "H": 9, "I": 22, "J": 13, "K": 11,
    "L": 12, "M": 18, "N": 20, "O": 11, "P": 9, "Q": 45, "R": 18, "S": 11,
    "T": 55, "U": 20,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w

flag_fill = PatternFill("solid", fgColor="FDECEC")
flagged_col_idx = list(df.columns).index("flagged") + 1
for row_idx in range(2, ws.max_row + 1):
    if ws.cell(row=row_idx, column=flagged_col_idx).value in (True, "True", "TRUE"):
        for col_idx in range(1, ws.max_column + 1):
            ws.cell(row=row_idx, column=col_idx).fill = flag_fill

wb.save(xlsx_path)
print(f"Wrote {xlsx_path} ({len(df)} rows, formatted, flagged rows highlighted)")

n_flagged = df["flagged"].sum()
print(f"\nSummary: {len(df)} total transactions, {n_flagged} flagged for review, "
      f"{(df['injected_fraud_pattern'] != '').sum()} known injected fraud patterns (for validation).")
